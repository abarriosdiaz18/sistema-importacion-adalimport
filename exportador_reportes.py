"""
exportador_reportes.py — ADALIMPORT
Genera reportes descargables (Excel .xlsx y PDF via HTML) del lote aprobado.

INSTRUCCIÓN DE INSTALACIÓN:
  Coloca este archivo en la misma carpeta que app_adalimport_streamlit.py.
  No requiere dependencias nuevas: usa openpyxl (ya presente en el proyecto).
"""
import io
from datetime import datetime


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def generar_excel_reporte(resultados: list, modo: str, lote_id: str,
                          costo_total: float, ganancia_total: float,
                          env_total: float) -> bytes:
    """
    Genera un .xlsx con dos hojas:
      • 'Detalle' — tabla de productos con todos los costos
      • 'Resumen'  — KPIs del lote (inversión, flete, ganancia, ROI)
    Retorna los bytes del archivo listo para st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter

    # ── Paleta ADALIMPORT ──────────────────────────────────────────────
    NAVY        = "0B1A2C"
    GOLD        = "C9A84C"
    GOLD_LIGHT  = "F5DFA0"
    GREEN_NEON  = "00E676"
    WHITE       = "E8E8E8"
    CELL_BG     = "0D1B2A"
    BORDER_COL  = "1E3A5F"

    thin  = Side(style="thin",   color=BORDER_COL)
    thick = Side(style="medium", color=GOLD)
    borde_data  = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════
    # HOJA 1 — DETALLE
    # ══════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Detalle"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # ── Encabezado principal ──────────────────────────────────────────
    ws.merge_cells("A1:Q1")
    c = ws["A1"]
    c.value = (f"REPORTE DE IMPORTACIÓN — ADALIMPORT   |   Lote: {lote_id}"
               f"   |   Vía: {'AÉREO' if modo == 'aereo' else 'MARÍTIMO'}"
               f"   |   {datetime.now().strftime('%d/%m/%Y')}")
    c.font      = Font(name="Arial", bold=True, color=GOLD, size=13)
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Cabeceras de columna ──────────────────────────────────────────
    cabeceras = [
        "Producto", "Tienda", "Cant.", "Costo Unit.",
        "Sales Tax", "Envío Int.", "Flete Courier",
        "Costo Real Unit.", "Costo Real Total",
        "Precio ML", "Ganancia Unit.", "Ganancia Total",
        "Margen %", "ROI %",
        "Peso (lb)", "Vol. (ft³)", "Modo Envío"
    ]
    col_widths = [32, 16, 7, 13, 11, 11, 14, 16, 16, 12, 14, 14, 10, 10, 11, 10, 14]

    ws.row_dimensions[2].height = 5   # spacer visual
    ws.row_dimensions[3].height = 22

    for col_idx, (header, width) in enumerate(zip(cabeceras, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = Font(name="Arial", bold=True, color=GOLD_LIGHT, size=9)
        cell.fill      = PatternFill("solid", fgColor="0D2240")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = borde_data
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Filas de datos ────────────────────────────────────────────────
    FMT_USD = '"$"#,##0.00'
    FMT_PCT = '0.0%'

    for row_num, r in enumerate(resultados, start=4):
        alt_fill = PatternFill("solid", fgColor=CELL_BG if row_num % 2 == 0 else "0A1520")
        gan_unit = r.get("ganancia_neta", r["precio_ml"] - r["costo_real"])
        gan_tot  = gan_unit * r["cantidad"]
        margen   = (gan_unit / r["precio_ml"]) if r["precio_ml"] > 0 else 0
        roi      = (gan_unit / r["costo_real"]) if r["costo_real"] > 0 else 0
        peso_lb  = r.get("peso_real", r.get("peso_lb", 0))
        vol_ft3  = r.get("vol_ft3", 0)

        fila = [
            r["nombre"],
            r.get("tienda", ""),
            r["cantidad"],
            r["precio_tienda"],
            r.get("sales_tax", 0),
            r.get("envio_interno", 0),
            r["envio_courier"],
            r["costo_real"],
            r["costo_real"] * r["cantidad"],
            r["precio_ml"],
            gan_unit,
            gan_tot,
            margen,
            roi,
            peso_lb,
            vol_ft3,
            "Aéreo" if modo == "aereo" else "Marítimo",
        ]

        for col_idx, val in enumerate(fila, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill      = alt_fill
            cell.border    = borde_data
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(name="Consolas", color=WHITE, size=9)

            if col_idx in (4, 5, 6, 7, 8, 9, 10, 11, 12):
                cell.number_format = FMT_USD
            elif col_idx in (13, 14):
                cell.number_format = FMT_PCT
                if isinstance(val, (int, float)):
                    cell.font = Font(name="Consolas", size=9,
                                     color=GREEN_NEON if val >= 0 else "FF1744")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name="Arial", color=WHITE, size=9, bold=True)

    # ── Fila de totales ───────────────────────────────────────────────
    total_row = 4 + len(resultados)
    ws.row_dimensions[total_row].height = 20
    ws.merge_cells(f"A{total_row}:H{total_row}")
    c_tot = ws[f"A{total_row}"]
    c_tot.value     = "TOTALES DEL LOTE"
    c_tot.font      = Font(name="Arial", bold=True, color=GOLD, size=10)
    c_tot.fill      = PatternFill("solid", fgColor="0D2240")
    c_tot.alignment = Alignment(horizontal="right", vertical="center")
    c_tot.border    = borde_data

    for col_idx, val in [(9, costo_total), (12, ganancia_total)]:
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font      = Font(name="Consolas", bold=True, color=GOLD, size=10)
        cell.fill      = PatternFill("solid", fgColor="0D2240")
        cell.number_format = FMT_USD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = borde_data

    roi_global = ganancia_total / costo_total if costo_total > 0 else 0
    c_roi = ws.cell(row=total_row, column=14, value=roi_global)
    c_roi.font         = Font(name="Consolas", bold=True, size=10,
                               color=GREEN_NEON if roi_global >= 0 else "FF1744")
    c_roi.fill         = PatternFill("solid", fgColor="0D2240")
    c_roi.number_format = FMT_PCT
    c_roi.alignment    = Alignment(horizontal="center", vertical="center")
    c_roi.border       = borde_data

    # ══════════════════════════════════════════════════════════════════
    # HOJA 2 — RESUMEN EJECUTIVO
    # ══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Resumen")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20

    def kpi_row(ws, row, label, value, fmt=FMT_USD, color=WHITE):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font      = Font(name="Arial", color="888888", size=10)
        lc.fill      = PatternFill("solid", fgColor=CELL_BG)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border    = borde_data

        vc = ws.cell(row=row, column=2, value=value)
        vc.font           = Font(name="Consolas", bold=True, color=color, size=11)
        vc.fill           = PatternFill("solid", fgColor=CELL_BG)
        vc.number_format  = fmt
        vc.alignment      = Alignment(horizontal="center", vertical="center")
        vc.border         = borde_data
        ws.row_dimensions[row].height = 22

    ws2.merge_cells("A1:B1")
    t = ws2["A1"]
    t.value     = "ADALIMPORT — RESUMEN EJECUTIVO"
    t.font      = Font(name="Arial", bold=True, color=GOLD, size=14)
    t.fill      = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32
    ws2.row_dimensions[2].height = 8

    kpi_row(ws2, 3,  "Lote ID",             lote_id,         fmt="@",  color=GOLD)
    kpi_row(ws2, 4,  "Vía de envío",        "AÉREO" if modo=="aereo" else "MARÍTIMO",
            fmt="@", color=GREEN_NEON)
    kpi_row(ws2, 5,  "Fecha del reporte",   datetime.now().strftime("%d/%m/%Y"),
            fmt="@", color=WHITE)
    kpi_row(ws2, 6,  "N° de productos",     len(resultados),  fmt="0",  color=WHITE)
    ws2.row_dimensions[7].height = 8
    kpi_row(ws2, 8,  "Flete Courier Total", env_total,        color=GREEN_NEON)
    kpi_row(ws2, 9,  "Inversión Total",     costo_total,      color=WHITE)
    kpi_row(ws2, 10, "Ganancia Estimada",   ganancia_total,
            color=GREEN_NEON if ganancia_total >= 0 else "FF1744")
    kpi_row(ws2, 11, "ROI del Lote",        roi_global,       fmt=FMT_PCT,
            color=GREEN_NEON if roi_global >= 0 else "FF1744")
    ws2.row_dimensions[12].height = 8

    mini_h = ws2.cell(row=13, column=1, value="Producto")
    mini_h.font  = Font(name="Arial", bold=True, color=GOLD_LIGHT, size=9)
    mini_h.fill  = PatternFill("solid", fgColor="0D2240")
    mini_h.border = borde_data
    mini_p = ws2.cell(row=13, column=2, value="Precio ML")
    mini_p.font  = Font(name="Arial", bold=True, color=GOLD_LIGHT, size=9)
    mini_p.fill  = PatternFill("solid", fgColor="0D2240")
    mini_p.border = borde_data

    for i, r in enumerate(resultados, start=14):
        c1 = ws2.cell(row=i, column=1, value=r["nombre"])
        c1.font      = Font(name="Arial", color=WHITE, size=9)
        c1.fill      = PatternFill("solid", fgColor=CELL_BG)
        c1.border    = borde_data
        c1.alignment = Alignment(horizontal="left")
        c2 = ws2.cell(row=i, column=2, value=r["precio_ml"])
        c2.font           = Font(name="Consolas", color=GREEN_NEON, size=9)
        c2.fill           = PatternFill("solid", fgColor=CELL_BG)
        c2.number_format  = FMT_USD
        c2.border         = borde_data
        c2.alignment      = Alignment(horizontal="center")
        ws2.row_dimensions[i].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — IMPORTACIÓN MASIVA WEB (Supabase)
# ══════════════════════════════════════════════════════════════════════════════

def _recortar_titulo(nombre: str, max_chars: int = 60) -> str:
    """
    Regla de Negocio — Títulos Web (Truncado Inteligente):
    Recorta el nombre del producto a un máximo de `max_chars` caracteres
    SIN cortar palabras a la mitad. Si el texto supera el límite, se corta
    en el último espacio disponible antes del límite, garantizando que todas
    las palabras queden completas.

    Ejemplos:
        "Auriculares Bluetooth Premium para deportes extremos" (52 chars) → sin cambios
        "Auriculares Bluetooth de Alta Fidelidad con Cancelación de Ruido Activa" →
        "Auriculares Bluetooth de Alta Fidelidad con Cancelación de" (59 chars)
    """
    nombre = nombre.strip()
    if len(nombre) <= max_chars:
        return nombre
    # Cortar en el límite y retroceder hasta el último espacio
    truncado = nombre[:max_chars]
    truncado = truncado.rsplit(' ', 1)[0]
    return truncado.strip()


def generar_excel_importacion_web(resultados: list, lote_id: str = "") -> bytes:
    """
    Genera un .xlsx plano (sin estilos) con la estructura exacta requerida
    para la importación masiva de productos en la plataforma web (Supabase).

    Columnas (en minúsculas, Fila 1) — orden exacto para Supabase:
        lote_id | title | description | stock | category | price | image

    Mapeo:
        lote_id     ← parámetro lote_id  (ej. "MAR-011")
        title       ← r["nombre"] truncado inteligente a 60 chars sin cortar palabras
        description ← r.get("descripcion", r.get("description", ""))
                      campo de descripción del ítem; string vacío si no existe
        stock       ← r["cantidad"]
        category    ← r.get("category", r.get("categoria", ""))
                      campo del Excel origen; vacío si no existe
        price       ← r.get("precio_ml_objetivo", r.get("precio_ml", 0))
        image       ← r.get("image_url", "")  — URL Supabase ya inyectada

    Columnas eliminadas (no van a la web):
        costo_real, costo_tienda, sales_tax, envio_interno,
        envio_courier, ganancia_neta, margen, roi, decision.

    Retorna los bytes del archivo listo para st.download_button.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # ── Fila 1: Encabezados — orden exacto Supabase ───────────────────────────
    encabezados = ["lote_id", "title", "description", "stock", "category", "price", "image"]
    ws.append(encabezados)

    # ── Filas de datos ────────────────────────────────────────────────────────
    for r in resultados:
        titulo   = _recortar_titulo(r.get("nombre", ""))
        precio   = r.get("precio_ml_objetivo", r.get("precio_ml", 0))
        # category viene del Excel original de productos si fue cargado;
        # se prueba "category" primero (nombre Supabase) y luego "categoria" (nombre interno)
        categoria = r.get("category", r.get("categoria", ""))
        imagen   = r.get("image_url") or r.get("image") or ""
        # Normalizar None/NaN a string vacío
        if not isinstance(imagen, str) or imagen.lower() in ("none", "nan"):
            imagen = ""
        # description: se prueba "descripcion" (nombre interno) y luego "description" (Supabase)
        descripcion = r.get("descripcion", r.get("description", ""))
        if not isinstance(descripcion, str):
            descripcion = ""

        fila = [
            lote_id,
            titulo,
            descripcion,
            r.get("cantidad", 0),
            categoria,
            precio,
            imagen,
        ]
        ws.append(fila)

    # ── Serialización en memoria ──────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PDF (via HTML auto-imprimible)
# ══════════════════════════════════════════════════════════════════════════════
def generar_pdf_reporte(resultados: list, modo: str, lote_id: str,
                        costo_total: float, ganancia_total: float,
                        env_total: float) -> bytes:
    """
    Genera un HTML profesional auto-imprimible como PDF.
    El navegador lo abre con un botón '🖨️ Imprimir / Guardar como PDF'.
    Retorna bytes de un archivo .html.
    """
    roi_global = (ganancia_total / costo_total * 100) if costo_total > 0 else 0
    via_label  = "AÉREO ✈️" if modo == "aereo" else "MARÍTIMO 🚢"
    entrega    = "3–5 días hábiles" if modo == "aereo" else "20–35 días hábiles"
    fecha      = datetime.now().strftime("%d/%m/%Y")

    # ── Tabla simplificada: 5 columnas — Resumen de Lote limpio ──────────────
    # Costo Unitario = precio_tienda + sales_tax (costo de compra con impuesto local)
    # Costo Real     = costo puesto en destino (precio_tienda + tax + envio_interno + flete)
    filas_html = ""
    for r in resultados:
        costo_unitario = r['precio_tienda'] + r.get('sales_tax', 0)
        filas_html += f"""
        <tr>
            <td style="text-align:left;font-weight:600">{r['nombre']}</td>
            <td>{r['cantidad']}</td>
            <td>${costo_unitario:.2f}</td>
            <td>${r['costo_real']:.2f}</td>
            <td>${r['precio_ml']:.2f}</td>
        </tr>"""

    color_gan  = "#00875A" if ganancia_total >= 0 else "#CC1A1A"
    color_roi  = "#00875A" if roi_global >= 0 else "#CC1A1A"
    gan_tot_td = f'style="color:{color_gan}"'
    roi_td     = f'style="color:{color_roi}"'

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Reporte ADALIMPORT — {lote_id}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Inter:wght@300;400;600;700&display=swap');
  *{{margin:0;padding:0;box-sizing:border-box}}
  body{{font-family:'Inter',sans-serif;background:#fff;color:#111827;padding:32px 40px}}
  .header{{background:linear-gradient(135deg,#0B1A2C,#122a4a);color:#C9A84C;
           padding:24px 32px;border-radius:12px;margin-bottom:24px;
           display:flex;justify-content:space-between;align-items:center;
           border:2px solid rgba(201,168,76,0.3)}}
  .header h1{{font-family:'Space Mono',monospace;font-size:1.4rem;letter-spacing:2px}}
  .header .meta{{font-size:0.75rem;color:#aaa;text-align:right;line-height:1.8}}
  .via-badge{{display:inline-block;background:rgba(0,230,118,0.15);color:#00875A;
              border:1px solid #00875A;padding:3px 12px;border-radius:20px;
              font-family:'Space Mono',monospace;font-size:0.8rem;font-weight:700;margin-top:4px}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:24px}}
  .kpi{{background:#f8f9fb;border:1px solid #e2e8f0;border-radius:10px;padding:14px;text-align:center}}
  .kpi .label{{font-size:0.62rem;color:#888;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:4px}}
  .kpi .value{{font-family:'Space Mono',monospace;font-size:1.25rem;font-weight:700;color:#0B1A2C}}
  .kpi .value.green{{color:#00875A}} .kpi .value.gold{{color:#B7862A}} .kpi .value.red{{color:#CC1A1A}}
  .section-title{{font-size:0.62rem;color:#888;text-transform:uppercase;letter-spacing:2px;
                  margin-bottom:8px;padding-bottom:5px;border-bottom:2px solid rgba(201,168,76,0.3)}}
  table{{width:100%;border-collapse:collapse;margin-bottom:24px;font-size:0.78rem}}
  thead th{{background:#0B1A2C;color:#C9A84C;padding:9px 7px;text-align:center;
            font-family:'Space Mono',monospace;font-size:0.65rem;letter-spacing:0.5px}}
  thead th:first-child{{text-align:left}}
  tbody tr:nth-child(even){{background:#f8f9fb}}
  tbody td{{padding:8px 7px;text-align:center;border-bottom:1px solid #e2e8f0;color:#374151}}
  tbody td:first-child{{text-align:left;color:#111827}}
  tfoot td{{background:#0B1A2C;color:#C9A84C;font-family:'Space Mono',monospace;
            font-weight:700;padding:9px 7px;text-align:center;font-size:0.82rem}}
  .footer{{text-align:center;font-size:0.62rem;color:#aaa;margin-top:24px;
           padding-top:12px;border-top:1px solid #e2e8f0;font-family:'Space Mono',monospace}}
  .print-btn{{background:#0B1A2C;color:#C9A84C;border:1px solid #C9A84C;
              padding:9px 24px;border-radius:8px;font-family:'Space Mono',monospace;
              font-size:0.82rem;cursor:pointer;margin-bottom:20px;letter-spacing:1px}}
  @media print{{.no-print{{display:none}}body{{padding:16px}}}}
</style>
</head>
<body>
<button class="print-btn no-print" onclick="window.print()">🖨️  IMPRIMIR / GUARDAR COMO PDF</button>

<div class="header">
  <div>
    <div style="font-size:0.65rem;letter-spacing:3px;color:#aaa;margin-bottom:4px">ADALIMPORT</div>
    <h1>REPORTE DE IMPORTACIÓN</h1>
    <div class="via-badge">{via_label}</div>
  </div>
  <div class="meta">
    <strong style="color:#C9A84C">Lote: {lote_id}</strong><br>
    {fecha}<br>
    Entrega estimada: {entrega}
  </div>
</div>

<div class="kpi-grid">
  <div class="kpi">
    <div class="label">Flete Courier</div>
    <div class="value gold">${env_total:.2f}</div>
  </div>
  <div class="kpi">
    <div class="label">Inversión Total</div>
    <div class="value">${costo_total:.2f}</div>
  </div>
  <div class="kpi">
    <div class="label">Ganancia Estimada</div>
    <div class="value {'green' if ganancia_total >= 0 else 'red'}">${ganancia_total:.2f}</div>
  </div>
  <div class="kpi">
    <div class="label">ROI del Lote</div>
    <div class="value {'green' if roi_global >= 0 else 'red'}">{roi_global:.1f}%</div>
  </div>
</div>

<div class="section-title">Resumen de Lote — Productos</div>
<table>
  <thead>
    <tr>
      <th>Producto</th><th>Cant.</th><th>Costo Unitario</th>
      <th>Costo Real</th><th>Precio ML</th>
    </tr>
  </thead>
  <tbody>{filas_html}</tbody>
  <tfoot>
    <tr>
      <td colspan="3" style="text-align:right">TOTAL INVERSIÓN →</td>
      <td>${costo_total:.2f}</td>
      <td>—</td>
    </tr>
  </tfoot>
</table>

<div class="footer">
  Generado por ADALIMPORT · Sistema de Gestión de Importaciones · {datetime.now().strftime("%d/%m/%Y %H:%M")}
</div>
</body>
</html>"""

    return html.encode("utf-8")
